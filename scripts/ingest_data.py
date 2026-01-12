import pandas as pd
import os
import json
import uuid
from openai import OpenAI
from dotenv import load_dotenv
import time
from openpyxl import load_workbook

# Resolve .env path relative to this script or CWD
# Resolve .env path relative to this script or CWD
BASE_DIR = os.getcwd()
ENV_PATH = os.path.join(BASE_DIR, 'web', '.env')
if not os.path.exists(ENV_PATH):
    ENV_PATH = os.path.join(BASE_DIR, '.env')
load_dotenv(ENV_PATH)

# Configuration
RESOURCES_DIR = "resources"
OUTPUT_DIR = "web/data"
JOURNALS_FILE = "List Scopus Outlet.xlsx"
SDGS_FILE = "SDGs Keyword.xlsx"
EMBEDDING_MODEL = "text-embedding-3-small"
BATCH_SIZE = 500

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

SDG_NAMES = {
    1: "No Poverty",
    2: "Zero Hunger",
    3: "Good Health and Well-being",
    4: "Quality Education",
    5: "Gender Equality",
    6: "Clean Water and Sanitation",
    7: "Affordable and Clean Energy",
    8: "Decent Work and Economic Growth",
    9: "Industry, Innovation and Infrastructure",
    10: "Reduced Inequalities",
    11: "Sustainable Cities and Communities",
    12: "Responsible Consumption and Production",
    13: "Climate Action",
    14: "Life Below Water",
    15: "Life on Land",
    16: "Peace, Justice and Strong Institutions",
    17: "Partnerships for the Goals"
}

def get_embeddings_batched_safe(text_map, model):
    """
    text_map: list of { index: int, text: string }
    Returns map of { index: embedding[] }
    """
    if not text_map:
        return {}
        
    results = {}
    total = len(text_map)
    print(f"Generating embeddings for {total} new items...")
    
    # Extract just texts for the batch
    all_indices = [item['index'] for item in text_map]
    all_texts = [item['text'].replace("\n", " ") for item in text_map]
    
    for i in range(0, total, BATCH_SIZE):
        batch_texts = all_texts[i:i+BATCH_SIZE]
        batch_indices = all_indices[i:i+BATCH_SIZE]
        
        try:
            response = client.embeddings.create(input=batch_texts, model=model)
            for j, data in enumerate(response.data):
                original_idx = batch_indices[j]
                results[original_idx] = data.embedding
            print(f"Processed {min(i+BATCH_SIZE, total)}/{total}")
        except Exception as e:
            print(f"Error in batch {i}: {e}")
            time.sleep(2) # Backoff
            
    return results

def ingest():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    # --- PROCESS JOURNALS (Main Cost Center) ---
    print("Processing Journals...")
    journals_jsonl_path = os.path.join(OUTPUT_DIR, "journals.jsonl")
    journals_json_path = os.path.join(OUTPUT_DIR, "journals.json")
    
    # 1. Load Existing Data from JSONL (Preferred for Resume)
    existing_names = set()
    
    if os.path.exists(journals_jsonl_path):
        print(f"Loading existing progress from {journals_jsonl_path}...", flush=True)
        try:
            with open(journals_jsonl_path, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        try:
                            # OPTIMIZATION: Only parse the name, don't keep full object in RAM
                            # We search for "name": "..." manually or use json.loads but drop result immediately
                            record = json.loads(line)
                            if 'name' in record:
                                existing_names.add(record['name'])
                        except: pass
            print(f"Loaded {len(existing_names)} existing records (Names only).", flush=True)
        except Exception as e:
            print(f"Error reading JSONL: {e}", flush=True)

    # Fallback checking is less critical now, assuming JSONL is primary


    # 1.5 Load ASJC Mapping
    asjc_map = {}
    asjc_path_file = os.path.join(RESOURCES_DIR, "ASJC1.xlsx")
    if os.path.exists(asjc_path_file):
        print("Loading ASJC mapping...", flush=True)
        df_asjc = pd.read_excel(asjc_path_file)
        for _, r in df_asjc.iterrows():
            code = str(r['Code']).strip()
            desc = str(r['Description']).strip() 
            asjc_map[code] = desc
        print("ASJC mapping loaded.", flush=True)

    # 2. Load Excel Data (with CSV Caching)
    journals_path = os.path.join(RESOURCES_DIR, JOURNALS_FILE)
    journals_cache_path = os.path.join(RESOURCES_DIR, "journals_cache.csv")
    
    if os.path.exists(journals_path):
        print(f"Loading Journal Data from {journals_path}...", flush=True)
        
        # Try loading from CSV cache first
        if os.path.exists(journals_cache_path):
            # Check modification times to ensure cache is fresh
            if os.path.getmtime(journals_cache_path) > os.path.getmtime(journals_path):
                print("Found cached CSV. Loading fast...", flush=True)
                try:
                    df_journals = pd.read_csv(journals_cache_path)
                    # Ensure strings
                    df_journals['All Science Journal Classification Codes (ASJC)'] = df_journals['All Science Journal Classification Codes (ASJC)'].astype(str)
                except:
                    print("Cache load failed, falling back to Excel.", flush=True)
                    df_journals = None
            else:
                print("Cache outdated. Reloading from Excel...", flush=True)
                df_journals = None
        else:
            df_journals = None

        # Fallback to Excel if no valid cache
        if df_journals is None:
            print("Reading Excel file (Streaming Mode - Low Memory)...", flush=True)
            
            # Use openpyxl in read-only mode to avoid loading everything into RAM
            wb = load_workbook(journals_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Identify columns
            headers = {}
            target_cols = {
                'Source Title': 'Source_Title',
                'scope': 'scope',
                'All Science Journal Classification Codes (ASJC)': 'All_Science_Journal_Classification_Codes_ASJC', 
                'Publisher': 'Publisher',
                'Sourcerecord ID': 'Sourcerecord_ID',
                'Active or Inactive': 'Active_or_Inactive',
                'Coverage': 'Coverage'
            }
            
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows)
            
            for idx, col_name in enumerate(header_row):
                if col_name in target_cols:
                    headers[idx] = target_cols[col_name]
            
            data_buffer = []
            count = 0
            print("Found columns, starting stream...", flush=True)

            class RowObject:
                pass

            # Generator to yield objects similar to itertuples
            def row_generator():
                nonlocal count
                for row_data in rows:
                    count += 1
                    if count % 5000 == 0:
                        print(f"Read {count} rows...", flush=True)
                        
                    obj = RowObject()
                    has_data = False
                    for col_idx, field_name in headers.items():
                        val = row_data[col_idx]
                        setattr(obj, field_name, val)
                        if field_name == 'Source_Title' and val:
                            has_data = True
                    
                    if has_data:
                        yield obj
            
            # For the first run, we stream directly into the processor
            # AND build the CSV cache simultaneously to avoid double reading
            print("Processing and building cache simultaneously...", flush=True)
            
            cache_data = [] # We still need to collect for CSV dump, but at least we process as we go? 
            # Actually, to save to CSV we need the full list or append mode. 
            # Let's just collect dicts for DataFrame at the end for caching, 
            # BUT yield to the main loop to show progress?
            # Complexity: Providing a generator to the next step matches the 'itertuples' interface.
            
            # Let's use a list for simplicity but fill it progressively? 
            # No, if memory is the issue, we should NOT build a huge list.
            # But we need to save the cache.
            # Compromise: Read all into a list of dicts (dict is heavier than tuple?)
            # Valid optimization: Just read what we need.
            
            iterator_source = row_generator()
            
        else:
            # If we loaded from CSV (df_journals is set), create an iterator from it
            print(f"Total rows to process: {len(df_journals)}", flush=True)
            iterator_source = df_journals.itertuples(index=False)

        # Use the unified iterator
        for row in iterator_source:
            # If coming from openpyxl, row is RowObject. If from pandas, it's namedtuple.
            # The attribute access logic below works for both!
            
            title = getattr(row, "Source_Title", "")
            # Basic cleanup if it's not a string (openpyxl might return None)
            if title is None: continue

            
            # If we already have it fully processed (with embedding), skip expensive logic
            if title in existing_names:
                # We do NOT append to final_list to save memory
                continue

            # ... Otherwise, prepare for processing ...
            # Access attributes safely. Column names are sanitized by Pandas:
            # 'scope' -> scope
            # 'All Science Journal Classification Codes (ASJC)' -> All_Science_Journal_Classification_Codes_ASJC
            # 'Publisher' -> Publisher
            # 'Sourcerecord ID' -> Sourcerecord_ID
            # 'Active or Inactive' -> Active_or_Inactive
            # 'Coverage' -> Coverage
            
            scope = str(getattr(row, "scope", ""))
            if scope == "nan": scope = ""
            
            # ASJC Optimization
            raw_asjc = str(getattr(row, "All_Science_Journal_Classification_Codes_ASJC", ""))
            if raw_asjc == "nan": raw_asjc = ""
            
            asjc_final = ""
            if raw_asjc:
                # Optimized single-pass
                parts = raw_asjc.replace(',', ';').split(';')
                # Generator expression for speed
                decoded_parts = (asjc_map.get(p.strip(), p.strip()) for p in parts if p.strip())
                asjc_final = "; ".join(decoded_parts)
            
            publisher = str(getattr(row, "Publisher", "Unknown Publisher"))
            if publisher == "nan": publisher = "Unknown Publisher"
            
            source_id = str(getattr(row, "Sourcerecord_ID", ""))
            if source_id == "nan": source_id = ""
            
            active_status = str(getattr(row, "Active_or_Inactive", "Unknown"))
            if active_status == "nan": active_status = "Unknown"
            
            years = str(getattr(row, "Coverage", ""))
            if years == "nan": years = ""
            
            coverage = f"{active_status} ({years})" if years else active_status
            link = f"https://www.scopus.com/sourceid/{source_id}" if source_id else ""

            if len(scope) > 4000:
                scope = scope[:4000] + "..."
            
            content = f"{title}. {scope}. matches ASJC codes: {asjc_final}".strip()
            
            # Reuse ID if possible
            rec_id = existing_journals[title]['id'] if title in existing_journals else str(uuid.uuid4())

            record = {
                "id": rec_id,
                "name": title,
                "scope": scope,
                "content": content,
                "publisher": publisher,
                "link": link,
                "coverage": coverage,
                "asjc": asjc_final
            }
            
            pending_records.append(record)

        print(f"Already processed: {len(final_list)}")
        print(f"Pending processing: {len(pending_records)}")
        
        # 3. Process Pending in Batches & SAVE INCREMENTALLY
        # Smaller batch size for better resume capability
        RESUME_BATCH_SIZE = 100 
        
        if os.getenv("OPENAI_API_KEY") and pending_records:
            total_pending = len(pending_records)
            for i in range(0, total_pending, RESUME_BATCH_SIZE):
                batch = pending_records[i : i + RESUME_BATCH_SIZE]
                
                # Prepare text map for embedding function
                to_embed_map = [{"index": idx, "text": item["content"]} for idx, item in enumerate(batch)]
                
                try:
                    # Call OpenAI
                    print(f"Embedding batch {i} to {i+len(batch)} of {total_pending}...")
                    embeddings = get_embeddings_batched_safe(to_embed_map, EMBEDDING_MODEL)
                    
                    # Assign embeddings
                    valid_batch = []
                    for idx, emb in embeddings.items():
                        batch[idx]["embedding"] = emb
                        valid_batch.append(batch[idx])
                    
                    # APPEND TO JSONL IMMEDIATELY (Checkpoint)
                    if valid_batch:
                        with open(journals_jsonl_path, "a", encoding="utf-8") as f:
                            for item in valid_batch:
                                f.write(json.dumps(item) + "\n")
                        print(f"Saved {len(valid_batch)} new records to {journals_jsonl_path} (Append-only)")
                        
                        # RAM OPTIMIZATION: Do not keep valid_batch in memory
                        # final_list.extend(valid_batch) 
                        
                except Exception as e:
                    print(f"CRITICAL ERROR in batch {i}: {e}", flush=True)
                    print("Waiting 10s before retry loop continues...", flush=True)
                    time.sleep(10)
        else:
            # No API key or no pending -> Just dump text data if needed
            if pending_records:
                print("No API Key or Dry Run: Saving pending records without embeddings.", flush=True)
                with open(journals_jsonl_path, "a", encoding="utf-8") as f:
                    for item in pending_records:
                        f.write(json.dumps(item) + "\n")

        # Memory Optimization: Skip syncing full JSON to avoid OOM
        # Users should use journals.jsonl
        # try:
        #     print("Syncing journals.json (Backup)...")
        #     with open(journals_json_path, "w", encoding="utf-8") as f:
        #         json.dump(final_list, f)
        #     print("Sync Complete.")
        # except Exception as e:
        #     print(f"Warning: Could not save full legacy JSON: {e}")

    else:
        print(f"File not found: {journals_path}")

    # --- PROCESS SDGs (Fast) ---
    # (Simpler logic here as it is small)
    print("Processing SDGs...")
    sdgs_path = os.path.join(RESOURCES_DIR, SDGS_FILE)
    if os.path.exists(sdgs_path):
        df_sdg = pd.read_excel(sdgs_path)
        sdg_data = []
        sdg_texts = []
        
        for _, row in df_sdg.iterrows():
            sdg_id = row['SDG']
            query = str(row['Query']) if pd.notna(row['Query']) else ""
            keywords = [k.strip().lower() for k in query.split(';') if k.strip()]
            name = SDG_NAMES.get(int(sdg_id), f"SDG {sdg_id}")
            
            text_for_embedding = f"{name}: {query}".strip()
            sdg_texts.append(text_for_embedding)

            sdg_data.append({
                "id": sdg_id,
                "name": name,
                "keywords": keywords
            })
        
        if os.getenv("OPENAI_API_KEY"):
            embeddings = get_embeddings_batched_safe([{'index': i, 'text': t} for i,t in enumerate(sdg_texts)], EMBEDDING_MODEL)
            for i, emb in embeddings.items():
                sdg_data[i]["embedding"] = emb
        
        with open(os.path.join(OUTPUT_DIR, "sdgs.json"), "w", encoding="utf-8") as f:
            json.dump(sdg_data, f)

if __name__ == "__main__":
    ingest()
